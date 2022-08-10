import openpyxl as xls
import tkinter.messagebox as mb
import os
import re

path = r"S:\00_Admin\Language_xml\Source xlsm\Language XML Data 2.0.xlsm"
target_file = r"S:\00_Admin\Language_xml\Language_import.sti"

wb_obj = xls.load_workbook(path)
sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = 36
list_of_keys = []
list_of_values = []
list_of_languages = []
dict_keys = {}

for j in range(1, column+1):
    list_of_languages.append(sheet_obj.cell(row=1, column=j).value)

for i in range(1,row):
    j_list = []
    for j in range(1, column+1):
        if re.search("gentext", str(sheet_obj.cell(row=i, column=j).value)):
            cell_obj = sheet_obj.cell(row=i, column=j)
            j_list.append(cell_obj.value.split('"')[3])
        elif re.search("gentext", str(sheet_obj.cell(row=i, column=1).value)):
            cell_obj = sheet_obj.cell(row=i, column=j)
            j_list.append("-")
    if not len(j_list) == 0: list_of_values.append(j_list)

for i in range(1, row):
    if re.search("gentext", str(sheet_obj.cell(row=i, column=1).value)):
        list_of_keys.append(sheet_obj.cell(row=i, column=1).value.split('"')[1])

i = 0
for key in list_of_keys:
    dict_keys[key] = list_of_values[i]
    i += 1

for lang in list_of_languages:
    try:
        with open(f"S:\\00_Admin\\Language_xml\\Language\\Language_{lang}.xml", "w", encoding="UTF-8") as target:
            target.writelines('<?xml version="1.0" encoding="UTF-8"?>\n')
            target.writelines('<l:langdata xmlns:l="http://www.schema.de/XSL/ST4DocuManagerlang">\n')
            target.writelines(f'\t<l:langblock xml:lang="{lang}">\n')
            for i in range(0, len(list_of_languages)):
                if i == list_of_languages.index(f'{lang}'):
                    for keys in list_of_keys:
                        key_value = dict_keys[keys][i]
                        target.writelines(f'\t\t<l:gentext key="{keys}" value="{key_value}"/>\n')
            target.writelines('\t</l:langblock>\n')
            target.writelines('</l:langdata>')
    except Exception as e:
        mb.showerror("Error when processing STI-File.", str(e))


with open(fr'{target_file}', 'w', encoding="UTF-8") as lang_xml:
    lang_xml.writelines('<stimport>\n')
    lang_xml.writelines(f'\t<node id="86781835" class="/class::BaseNodeClass/DocuManagerBaseClass/Resource'
                          f'/DocumentResource" parent="46764043">\n')
    for lang in list_of_languages:
        lang_xml.writelines(f'\t\t<attribute name="Resource" type="resource" aspect="{lang}">Language\Language_{lang}.xml</attribute>\n')
        lang_xml.writelines(f'\t\t<attribute name="Title" type="string" aspect="{lang}">Language\Language_{lang}.xml</attribute>\n')
    lang_xml.writelines(f'\t</node>\n')
    lang_xml.writelines(f'</stimport>\n')

mb.showinfo("Successful", "Programm ran successfully.")

